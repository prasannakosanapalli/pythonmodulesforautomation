# -*- coding: utf-8 -*-

import openpyxl
import openpyxl.cell import get_colomn_letter


a= openpyxl.load_workbook("path or variable")
a.get_active_sheet()
a.get_sheet_names()
a.get_sheet_by_name("name")
a.get_active_sheet()
a.title()
a.max_row
a.max_coloum
wb.save("path to save")
a.create_sheet("name")
a.remove_sheet("name")
a.create_sheet(index=,title="")

b=a.get_sheet_by_name("")
c=b.sheet.cell(row='',colomn='')
c.value()
sheet["a1":'c2']
b.row
b.colomn
b.cordinate
a.get_highest_row
a.get_highest_colmn
get_colomn_letter(26)
get_row_letter(2)
row_index_from_string('a')
formulas
sheet['b1']= 100
sheet["c1"]= 200
sheet['d1']= "=sum[b1:c1]"
cheet[d1].value()
